import streamlit as st
import pandas as pd
import requests
try:
    import cloudscraper
except ImportError:
    cloudscraper = None
from bs4 import BeautifulSoup
import re
import json
import logging
import sqlite3
import time
import io
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)
st.set_page_config(
    page_title="Monitor de Precios",
    page_icon="logo_form.jpg",
    layout="wide",
    initial_sidebar_state="collapsed"
)
RUTA_EXCEL      = "base_precios.xlsx"
import threading
import pickle
import os
_playwright_sem = threading.Semaphore(3)
DB_PATH         = "historial_precios.db"
import base64

# ── AUTENTICACIÓN ─────────────────────────────────────────────────────────────
import hashlib

_PASSWORD_HASH = hashlib.sha256("#Edge2026".encode()).hexdigest()

def _check_login(email: str, password: str) -> str | None:
    if not email.lower().endswith("@form.cl"):
        return "Credenciales inválidas"
    if hashlib.sha256(password.encode()).hexdigest() != _PASSWORD_HASH:
        return "Credenciales inválidas"
    return None

if not st.session_state.get("autenticado"):
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500;600&display=swap');
    .stApp { background: #f7f7f5 !important; font-family: 'DM Sans', sans-serif; }
    </style>
    <div style="max-width:380px;margin:80px auto 0">
        <div style="font-family:'Syne',sans-serif;font-size:1.6rem;font-weight:800;
                    color:#111111;margin-bottom:4px;text-align:center">
            Monitor de Precios</div>
        <div style="font-size:0.8rem;color:#5a5a5a;margin-bottom:28px;text-align:center">
            Form Design · Acceso restringido</div>
    </div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        email    = st.text_input("Correo @form.cl", placeholder="nombre@form.cl")
        password = st.text_input("Contraseña", type="password")
        submit   = st.form_submit_button("Ingresar", use_container_width=True)

    if submit:
        error = _check_login(email.strip(), password)
        if error:
            st.error(error)
        else:
            st.session_state["autenticado"] = True
            st.session_state["usuario_email"] = email.strip().lower()
            st.rerun()

    st.stop()

# ── LOGOUT (sidebar) ──────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"👤 `{st.session_state.get('usuario_email', '')}`")
    st.divider()
    if st.button("Cerrar sesión", use_container_width=True):
        st.session_state["autenticado"] = False
        st.session_state["usuario_email"] = ""
        st.rerun()

# ── SUPABASE CLIENT ─────────────────────────────────────────────────────────
def _get_supabase():
    try:
        from supabase import create_client
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception as e:
        logger.warning(f"[Supabase] No se pudo conectar: {e}")
        return None

# ── MODO VISTA ─────────────────────────────────────────────────────────────
_qp_raw = st.query_params
MODO_VISTA = (
    _qp_raw.get("modo", "").lower() == "vista"
    or os.environ.get("MODO_VISTA", "").strip() == "1"
)
MIN_PRECIO      = 1_000
MAX_PRECIO      = 10_000_000
MAX_WORKERS     = 8
REQUEST_TIMEOUT = 20
CACHE_TTL       = 600
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    "Accept-Language": "es-CL,es;q=0.9"
}

SSL_IGNORE_DOMAINS = [
    "interlight.cl",
    "tienda.contatto.cl",
    "vetaindomita.cl",
    "form.cl",
    "lasilleria.cl",
    "tolixcenter.cl",
    "solosillas.cl",
    "epicadeco.cl",
    "simplegracia.cl",
    "lablanqueria.cl",
    "sodimac.cl",
]

# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
:root {
    --yellow:      #F5D000;
    --yellow-dim:  #c9aa00;
    --yellow-bg:   #fdf9e3;
    --black:       #111111;
    --black-soft:  #1c1c1c;
    --gray-dark:   #2e2e2e;
    --gray-mid:    #5a5a5a;
    --gray-light:  #c0c0c0;
    --gray-line:   #e4e4e4;
    --white:       #ffffff;
    --bg:          #f7f7f5;
    --surface:     #ffffff;
    --surface2:    #f2f2f0;
    --green:       #1a9e6e;
    --red:         #d93a3a;
}
*, *::before, *::after { box-sizing: border-box; }
.stApp { background: var(--bg) !important; color: var(--black); font-family: 'DM Sans', sans-serif; }
[data-testid="stAppViewContainer"] { background: var(--bg) !important; }
[data-testid="stHeader"]            { background: var(--white) !important; border-bottom: 1px solid var(--gray-line); }
section[data-testid="stSidebar"]    { background: var(--white) !important; }
.header-wrap {
    display: flex; align-items: center; justify-content: space-between;
    padding: 24px 0 18px; border-bottom: 2px solid var(--yellow); margin-bottom: 28px;
}
.header-left { display: flex; align-items: center; gap: 16px; }
.header-accent { width: 6px; height: 42px; background: var(--yellow); border-radius: 3px; flex-shrink: 0; }
.header-title { font-family: 'Syne', sans-serif; font-size: 1.55rem; font-weight: 800; color: var(--black); letter-spacing: -0.4px; }
.header-sub { font-size: 0.77rem; color: var(--gray-mid); margin-top: 2px; }
.header-badge { background: var(--yellow); color: var(--black); font-size: 0.71rem; font-weight: 700; padding: 4px 12px; border-radius: 20px; font-family: 'DM Mono', monospace; }
.metrics-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin-bottom: 24px; }
.metric-card { background: var(--surface); border: 1px solid var(--gray-line); border-radius: 14px; padding: 18px 20px; position: relative; overflow: hidden; }
.metric-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: var(--accent, var(--gray-line)); }
.metric-card.yellow { --accent: var(--yellow); }
.metric-card.red    { --accent: var(--red); }
.metric-card.green  { --accent: var(--green); }
.metric-card.gray   { --accent: var(--gray-light); }
.metric-label { font-size: 0.67rem; color: var(--gray-mid); text-transform: uppercase; letter-spacing: 1px; font-weight: 600; margin-bottom: 8px; }
.metric-value { font-family: 'Syne', sans-serif; font-size: 2rem; font-weight: 800; color: var(--black); line-height: 1; }
.metric-sub { font-size: 0.71rem; color: var(--gray-light); margin-top: 6px; }
.product-name { font-weight: 500; color: var(--black-soft); max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; display: block; }
.product-name a { color: inherit; text-decoration: none; }
.product-name a:hover { color: var(--yellow-dim); text-decoration: underline; }
.price-pill { display: inline-block; background: var(--surface2); border: 1px solid var(--gray-line); border-radius: 6px; padding: 4px 10px; font-family: 'DM Mono', monospace; font-size: 0.8rem; color: var(--gray-dark); white-space: nowrap; }
.price-our { background: var(--yellow); border-color: var(--yellow-dim); color: var(--black); font-weight: 700; }
.empresa-tag { display: inline-block; background: var(--black); color: var(--white); font-size: 0.67rem; font-weight: 700; padding: 3px 7px; border-radius: 4px; text-transform: uppercase; }
.rubro-tag { display: inline-block; background: var(--yellow-bg); border: 1px solid var(--yellow); color: var(--black); font-size: 0.68rem; font-weight: 600; padding: 3px 8px; border-radius: 4px; }
.dif-mas   { color: var(--red);   font-weight: 700; font-family: 'DM Mono', monospace; font-size: 0.8rem; }
.dif-menos { color: var(--green); font-weight: 700; font-family: 'DM Mono', monospace; font-size: 0.8rem; }
.dif-igual { color: var(--gray-mid); font-family: 'DM Mono', monospace; font-size: 0.8rem; }
.price-up   { color: var(--red);   font-size: 0.72rem; font-weight: 600; }
.price-down { color: var(--green); font-size: 0.72rem; font-weight: 600; }
.status-none { color: var(--gray-mid); font-size: 0.72rem; font-style: italic; }
.img-thumb { border-radius: 7px; object-fit: cover; border: 1px solid var(--gray-line); }
.modal-title { font-family: 'Syne', sans-serif; font-size: 1.3rem; font-weight: 800; color: var(--black); margin-bottom: 4px; }
.modal-sub { color: var(--gray-mid); font-size: 0.8rem; margin-bottom: 24px; }
.kpi-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-bottom: 20px; }
.kpi-card { background: var(--surface2); border: 1px solid var(--gray-line); border-radius: 12px; padding: 15px 17px; }
.kpi-label { font-size: 0.67rem; color: var(--gray-mid); text-transform: uppercase; letter-spacing: 0.8px; margin-bottom: 6px; font-weight: 600; }
.kpi-val   { font-family: 'Syne', sans-serif; font-size: 1.35rem; font-weight: 800; color: var(--black); }
.kpi-sub   { font-size: 0.71rem; color: var(--gray-light); margin-top: 4px; }
.rival-row { display: flex; align-items: center; justify-content: space-between; padding: 10px 14px; border-radius: 10px; background: var(--surface2); border: 1px solid var(--gray-line); margin-bottom: 8px; font-size: 0.83rem; }
.rival-name  { color: var(--black-soft); font-weight: 500; }
.rival-price { font-family: 'DM Mono', monospace; color: var(--gray-dark); }
.section-title { font-family: 'Syne', sans-serif; font-size: 0.82rem; font-weight: 800; color: var(--black); text-transform: uppercase; letter-spacing: 1px; margin: 22px 0 12px; padding-bottom: 8px; border-bottom: 2px solid var(--yellow); }
.stButton > button {
    background: var(--black) !important; color: var(--yellow) !important;
    border: 2px solid var(--black) !important; border-radius: 8px !important;
    padding: 9px 22px !important; font-weight: 700 !important; font-size: 0.87rem !important;
    box-shadow: none !important; transition: all 0.15s !important;
}
.stButton > button:hover { background: var(--yellow) !important; color: var(--black) !important; border-color: var(--yellow) !important; }
.stDownloadButton > button {
    background: var(--yellow) !important; color: var(--black) !important;
    border: 2px solid var(--yellow-dim) !important; border-radius: 8px !important;
    padding: 9px 22px !important; font-weight: 700 !important; font-size: 0.87rem !important;
    box-shadow: none !important; transition: all 0.15s !important;
}
.stDownloadButton > button:hover { background: var(--yellow-dim) !important; border-color: var(--black) !important; }
div[data-testid="stButton"] > button { font-family: 'DM Mono', monospace !important; font-size: 0.72rem !important; padding: 3px 8px !important; border-radius: 5px !important; min-height: 0 !important; height: auto !important; line-height: 1.5 !important; }
.stSelectbox label, .stMultiSelect label { color: var(--gray-mid) !important; font-size: 0.71rem !important; text-transform: uppercase !important; letter-spacing: 0.8px !important; font-weight: 600 !important; }
[data-baseweb="select"] > div { background: var(--white) !important; border-color: var(--gray-line) !important; color: var(--black) !important; }
[data-baseweb="select"] svg { color: var(--black) !important; }
.stProgress > div > div { background: var(--yellow) !important; }
.empty-state { text-align: center; padding: 80px 0; }
.empty-icon  { font-size: 2.8rem; margin-bottom: 14px; }
.empty-text  { color: var(--gray-mid); font-size: 0.95rem; }
.empty-text strong { color: var(--black); }
.reporte-banner {
    display: flex; align-items: center; background: var(--yellow-bg);
    border: 1.5px solid var(--yellow); border-radius: 12px;
    padding: 14px 20px; margin-bottom: 20px; gap: 14px;
}
.reporte-icon  { font-size: 1.7rem; flex-shrink: 0; }
.reporte-title { font-family: 'Syne', sans-serif; font-weight: 800; font-size: 0.95rem; color: var(--black); }
.reporte-sub   { font-size: 0.73rem; color: var(--gray-mid); margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# BASE DE DATOS — HISTORIAL
# ─────────────────────────────────────────────
def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS historial (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL, sku TEXT NOT NULL, url TEXT NOT NULL,
            empresa TEXT, precio INTEGER, estado TEXT
        )
    """)
    con.commit(); con.close()
def guardar_precios(filas: list[dict]):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    con = sqlite3.connect(DB_PATH)
    con.executemany(
        "INSERT INTO historial (ts, sku, url, empresa, precio, estado) VALUES (?,?,?,?,?,?)",
        [(ts, f["sku"], f["url"], f["empresa"], f["precio"], f["estado"]) for f in filas]
    )
    con.commit(); con.close()
def precio_anterior(sku: str, url: str) -> Optional[int]:
    con = sqlite3.connect(DB_PATH)
    cur = con.execute(
        "SELECT precio FROM historial WHERE sku=? AND url=? AND precio IS NOT NULL ORDER BY ts DESC LIMIT 2",
        (sku, url)
    )
    rows = cur.fetchall(); con.close()
    return rows[1][0] if len(rows) >= 2 else None
def historial_sku(sku: str) -> pd.DataFrame:
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query(
        "SELECT ts, empresa, url, precio FROM historial WHERE sku=? AND precio IS NOT NULL ORDER BY ts DESC LIMIT 60",
        con, params=(sku,)
    )
    con.close(); return df
init_db()

# ── SNAPSHOT helpers (Supabase) ─────────────────────────────────────────────
def guardar_snapshot(df_final: pd.DataFrame, timestamp: str):
    try:
        data_b64 = base64.b64encode(pickle.dumps({"df_final": df_final, "timestamp": timestamp})).decode()
        sb = _get_supabase()
        if sb:
            sb.table("snapshots").delete().neq("id", 0).execute()
            sb.table("snapshots").insert({"ts": timestamp, "data": data_b64}).execute()
            logger.info("[snapshot] Guardado en Supabase")
        else:
            logger.warning("[snapshot] Sin conexión a Supabase")
    except Exception as e:
        logger.warning(f"[snapshot] No se pudo guardar: {e}")

def cargar_snapshot() -> tuple:
    try:
        sb = _get_supabase()
        if not sb:
            return None, None
        res = sb.table("snapshots").select("ts,data").order("created_at", desc=True).limit(1).execute()
        if not res.data:
            return None, None
        row = res.data[0]
        payload = pickle.loads(base64.b64decode(row["data"]))
        return payload.get("df_final"), payload.get("timestamp", "")
    except Exception as e:
        logger.warning(f"[snapshot] No se pudo cargar: {e}")
        return None, None

# ── PRECIOS EDITADOS en Supabase (persistentes) ────────────────────────────
def _cargar_precios_editados_sb() -> dict:
    try:
        sb = _get_supabase()
        if not sb:
            return {}
        res = sb.table("precios_editados").select("sku,empresa,precio").execute()
        if not res.data:
            return {}
        return {(r["sku"], r["empresa"]): r["precio"] for r in res.data}
    except Exception as e:
        logger.warning(f"[precios_editados] No se pudo cargar: {e}")
        return {}

def _guardar_precio_editado_sb(sku: str, empresa: str, precio: int):
    try:
        sb = _get_supabase()
        if not sb:
            return
        sb.table("precios_editados").upsert(
            {"sku": sku, "empresa": empresa, "precio": precio},
            on_conflict="sku,empresa"
        ).execute()
        logger.info(f"[precios_editados] Guardado: {sku}/{empresa} = ${precio}")
    except Exception as e:
        logger.warning(f"[precios_editados] No se pudo guardar: {e}")

def _eliminar_precio_editado_sb(sku: str, empresa: str):
    try:
        sb = _get_supabase()
        if not sb:
            return
        sb.table("precios_editados").delete().eq("sku", sku).eq("empresa", empresa).execute()
        logger.info(f"[precios_editados] Eliminado: {sku}/{empresa}")
    except Exception as e:
        logger.warning(f"[precios_editados] No se pudo eliminar: {e}")

def _limpiar_todos_precios_editados_sb():
    try:
        sb = _get_supabase()
        if not sb:
            return
        sb.table("precios_editados").delete().neq("id", 0).execute()
        logger.info("[precios_editados] Todos eliminados")
    except Exception as e:
        logger.warning(f"[precios_editados] No se pudo limpiar: {e}")

# ─────────────────────────────────────────────
# DATACLASS
# ─────────────────────────────────────────────
@dataclass
class ResultadoScrape:
    url:        str
    nombre:     Optional[str] = None
    precio_txt: Optional[str] = None
    precio_num: Optional[int] = None
    imagen:     Optional[str] = None
    estado:     str           = "ok"
    empresa:    str           = ""

# ─────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────
def limpiar_texto(texto) -> str:
    if not texto: return ""
    return re.sub(r"\s+", " ", str(texto).replace("\n", " ")).strip()
def limpiar_precio(texto) -> Optional[int]:
    if texto is None: return None
    if isinstance(texto, (int, float)):
        v = int(round(float(texto)))
        return v if MIN_PRECIO <= v <= MAX_PRECIO else None
    texto = str(texto).replace("$", "").replace(" ", "").strip()
    if re.match(r"^\d+\.\d{1,2}$", texto):
        v = int(round(float(texto)))
        return v if MIN_PRECIO <= v <= MAX_PRECIO else None
    texto = texto.replace(".", "").replace(",", "")
    return int(texto) if texto.isdigit() and MIN_PRECIO <= int(texto) <= MAX_PRECIO else None
def fmt(valor: int) -> str:
    return f"{valor:,.0f}".replace(",", ".")
def get_empresa(url: str) -> str:
    d = urlparse(url).netloc.lower().replace("www.", "")
    if d.startswith("sodimac."): return "Sodimac"
    if d.startswith("tottus."):  return "Tottus"
    if d.startswith("simple."):  return d.split(".")[1].capitalize()
    p = d.split(".")
    return (p[-2] if len(p) >= 2 else d).capitalize()

# ─────────────────────────────────────────────
# Request unificado con manejo SSL + retry
# ─────────────────────────────────────────────
def _hacer_request(url: str):
    no_verify = any(d in url for d in SSL_IGNORE_DOMAINS)
    if any(x in url for x in ["ripley.cl", "falabella.com"]):
        s = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        return s.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=not no_verify)
    if "sodimac.falabella.com" in url or "sodimac.cl" in url:
        try:
            return requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=not no_verify)
        except requests.exceptions.SSLError:
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            logger.warning(f"[SSL retry] {url[:70]}")
            return requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=False)
    if no_verify:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=False)
    try:
        return requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
    except requests.exceptions.SSLError:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        logger.warning(f"[SSL retry automático] {url[:70]}")
        return requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=False)

# ─────────────────────────────────────────────
# ESTRATEGIAS SCRAPING
# ─────────────────────────────────────────────
def _try_jsonld(soup, **_) -> Optional[int]:
    for s in soup.find_all("script", type="application/ld+json"):
        try:
            if not s.string: continue
            data = json.loads(s.string)
            items = data if isinstance(data, list) else [data]
            for item in items:
                if not isinstance(item, dict) or item.get("@type") != "Product": continue
                offers = item.get("offers", {})
                cands = []
                if isinstance(offers, dict):
                    for k in ("price", "lowPrice"):
                        p = limpiar_precio(offers.get(k))
                        if p: cands.append(p)
                elif isinstance(offers, list):
                    for o in offers:
                        p = limpiar_precio(o.get("price"))
                        if p: cands.append(p)
                if cands: return min(cands)
        except Exception as e:
            logger.debug(f"[jsonld] {e}")
    return None
def _try_woocommerce(soup, **_) -> Optional[int]:
    try:
        c = soup.find(class_="price")
        if not c: return None
        cands = []
        for span in c.find_all("span", class_="woocommerce-Price-amount"):
            if span.find_parent("del"): continue
            p = limpiar_precio(limpiar_texto(span.get_text()))
            if p: cands.append(p)
        return min(cands) if cands else None
    except Exception as e:
        logger.debug(f"[woo] {e}"); return None
def _try_shopify(soup, **_) -> Optional[int]:
    try:
        for s in soup.find_all("script"):
            if s.string and "ShopifyAnalytics" in s.string:
                m = re.search(r'"price":\s*"(\d+)"', s.string)
                if m: return limpiar_precio(m.group(1))
    except Exception as e:
        logger.debug(f"[shopify] {e}")
    return None
def _try_meta(soup, **_) -> Optional[int]:
    try:
        meta = soup.find("meta", property="product:price:amount")
        if meta and meta.get("content"): return limpiar_precio(meta["content"])
    except: pass
    return None
def _try_easy(html, url, **_) -> Optional[int]:
    if "easy.cl" not in url: return None
    try:
        m = re.search(r'"prices"\s*:\s*\{([^}]+)\}', html)
        if not m: return None
        cands = []
        for k in ("offerPrice", "normalPrice"):
            mm = re.search(rf'"{k}"\s*:\s*(\d+)', m.group(1))
            if mm:
                p = limpiar_precio(mm.group(1))
                if p: cands.append(p)
        return min(cands) if cands else None
    except: return None
def _try_magento(soup, **_) -> Optional[int]:
    try:
        cands = [limpiar_precio(t.get("data-price-amount"))
                 for t in soup.find_all(attrs={"data-price-amount": True})]
        cands = [p for p in cands if p]
        return min(cands) if cands else None
    except: return None
def _try_fallback(soup, **_) -> Optional[int]:
    try:
        cands = [limpiar_precio(e) for e in re.findall(r'\$\s?\d[\d\.\,]*', soup.get_text(" "))]
        cands = sorted([p for p in cands if p])
        return cands[len(cands)//2] if cands else None
    except: return None
ESTRATEGIAS = [
    ("JSON-LD",     _try_jsonld),
    ("WooCommerce", _try_woocommerce),
    ("Shopify",     _try_shopify),
    ("Meta",        _try_meta),
    ("Easy",        _try_easy),
    ("Magento",     _try_magento),
    ("Fallback",    _try_fallback),
]

PLAYWRIGHT_DOMAINS = [
    "mercadolibre.cl",
    "zara.com",
    "zarahome.com",
    "falabella.com",
]

_PW_SCRIPT = r"""
import sys, json, time, re
from playwright.sync_api import sync_playwright

url       = sys.argv[1]
site_hint = sys.argv[2]

def limpiar(txt, es_zara=False, es_zarahome=False):
    if txt is None: return None
    s = str(txt).replace("$","").replace(" ","").replace("\xa0","").strip()
    try:
        if re.match(r"^\d+\.\d+$", s):
            v = int(round(float(s)))
        else:
            v = int(s.replace(".","").replace(",",""))
        return v if 1000 <= v <= 10_000_000 else None
    except:
        return None

es_zara      = site_hint == "zara"
es_zarahome  = site_hint == "zarahome"
es_falabella = site_hint == "falabella"
es_shopify   = site_hint == "shopify"
es_zara_any  = es_zara or es_zarahome

try:
    with sync_playwright() as pw:
        if es_zara_any:
            browser = pw.firefox.launch(headless=True)
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
                locale="es-CL", viewport={"width":1280,"height":900}
            )
        else:
            browser = pw.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled","--no-sandbox",
                      "--disable-dev-shm-usage","--disable-gpu"]
            )
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
                locale="es-CL", viewport={"width":1280,"height":800}
            )
            ctx.add_init_script(
                "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
                "Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3]});"
                "window.chrome={runtime:{}};"
            )

        page = ctx.new_page()

        if es_falabella:
            try:
                page.goto(url, timeout=35000, wait_until="domcontentloaded")
            except: pass
            try:
                page.wait_for_selector(
                    "[class*='price'], [data-testid*='price'], [class*='Price'], [class*='prices-0'], meta[property='product:price:amount']",
                    timeout=10000
                )
            except: pass
            time.sleep(1.5)
        elif es_zara_any:
            page.goto(url, timeout=50000, wait_until="domcontentloaded")
            time.sleep(4)
            from urllib.parse import parse_qs, urlparse as _up
            _qs = parse_qs(_up(url).query)
            _pelement = _qs.get("pelement", [None])[0]
            if _pelement:
                try:
                    page.click(f"[data-product-element-id='{_pelement}']", timeout=3000)
                    time.sleep(1.5)
                except: pass
        elif es_shopify:
            page.goto(url, timeout=45000, wait_until="networkidle")
            try:
                page.wait_for_selector(
                    "[class*='price'], [class*='Price'], meta[property='product:price:amount']",
                    timeout=10000
                )
            except: pass
            time.sleep(2)
        else:
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            try:
                page.wait_for_function(
                    "() => Array.from(document.querySelectorAll('script[type=\"application/ld+json\"]')).some(s => s.textContent.includes('price'))",
                    timeout=12000
                )
            except: pass
            time.sleep(1.5)

        html   = page.content()
        nombre = None
        imagen = None
        precio = None

        try:
            meta_el = page.query_selector("meta[property='product:price:amount']")
            if meta_el:
                content_val = meta_el.get_attribute("content")
                if content_val:
                    precio = limpiar(content_val, es_zara=es_zara, es_zarahome=es_zarahome)
        except: pass

        if not precio:
            from bs4 import BeautifulSoup
            import json as _json
            soup = BeautifulSoup(html, "html.parser")
            cands = []
            for tag in soup.find_all("script", type="application/ld+json"):
                raw = tag.string or tag.get_text()
                if not raw or "price" not in raw.lower(): continue
                try: data = _json.loads(raw)
                except: continue
                items = data if isinstance(data, list) else [data]
                for item in items:
                    if not isinstance(item, dict): continue
                    offers = item.get("offers") or item.get("Offers")
                    if not offers: continue
                    for o in (offers if isinstance(offers, list) else [offers]):
                        if not isinstance(o, dict): continue
                        for k in ("price","lowPrice"):
                            p = limpiar(o.get(k), es_zara=es_zara, es_zarahome=es_zarahome)
                            if p: cands.append(p)
            if cands: precio = min(cands)

        if not precio and es_zara_any:
            for sel in ["[class*='price'] [class*='amount']","[class*='price-current']","[class*='money-amount']","span[class*='price']","[data-testid='price']","[class*='price-item']"]:
                try:
                    el = page.query_selector(sel)
                    if el:
                        p = limpiar(el.inner_text().strip(), es_zara=es_zara, es_zarahome=es_zarahome)
                        if p: precio = p; break
                except: pass

        if not precio and es_zara_any:
            for pattern in [r'"price"\s*:\s*"?([\d\.]+)"?', r'"amount"\s*:\s*"?([\d\.]+)"?', r'CLP\s*([\d\.]+)']:
                m = re.search(pattern, html)
                if m:
                    p = limpiar(m.group(1), es_zara=es_zara, es_zarahome=es_zarahome)
                    if p: precio = p; break

        if not precio and es_falabella:
            next_data_match = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
            if next_data_match:
                try:
                    nd_str = next_data_match.group(1)
                    for pattern in [r'"normalPrice"\s*:\s*(\d+)',r'"offerPrice"\s*:\s*(\d+)',r'"currentPrice"\s*:\s*(\d+)',r'"salePrice"\s*:\s*(\d+)',r'"price"\s*:\s*"?(\d{4,8})"?']:
                        m = re.search(pattern, nd_str)
                        if m:
                            p = limpiar(m.group(1))
                            if p: precio = p; break
                except: pass

        if not precio and es_falabella:
            for sel in ["[class*='price-box'] [class*='prices-0']","[class*='price-box__prices']","li[class*='prices-0']","span[class*='copy10']","[data-testid='price-without-discount']","[data-testid='price']","[class*='Price']"]:
                try:
                    el = page.query_selector(sel)
                    if el:
                        p = limpiar(el.inner_text().strip())
                        if p: precio = p; break
                except: pass

        if not precio and es_falabella:
            cands_regex = []
            for pattern in [r'"normalPrice"\s*:\s*(\d{4,8})',r'"offerPrice"\s*:\s*(\d{4,8})',r'"currentPrice"\s*:\s*(\d{4,8})',r'"salePrice"\s*:\s*(\d{4,8})',r'"price"\s*:\s*"?(\d{4,8})"?',r':\"(\d{4,8})\",\"currency\":\"CLP\"',r'"listPrice"\s*:\s*(\d{4,8})',r'"sellingPrice"\s*:\s*(\d{4,8})']:
                for m in re.finditer(pattern, html):
                    p = limpiar(m.group(1))
                    if p: cands_regex.append(p)
            if cands_regex:
                from collections import Counter
                precio = Counter(cands_regex).most_common(1)[0][0]

        if not precio:
            try:
                el = page.query_selector("[data-price]")
                if el: precio = limpiar(el.get_attribute("data-price"))
            except: pass

        try:
            og = page.query_selector("meta[property='og:title']")
            nombre = og.get_attribute("content") if og else page.title()
        except: pass
        try:
            og2 = page.query_selector("meta[property='og:image']")
            imagen = og2.get_attribute("content") if og2 else None
        except: pass

        browser.close()

    print(json.dumps({"precio": precio, "nombre": nombre, "imagen": imagen}))

except Exception as e:
    print(json.dumps({"precio": None, "nombre": None, "imagen": None, "error": str(e)}))
"""

def _scrape_playwright(url: str, empresa: str) -> ResultadoScrape:
    import subprocess, sys, json as _json, tempfile, os
    try:
        with _playwright_sem:
            with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False, encoding="utf-8") as tf:
                tf.write(_PW_SCRIPT)
                script_path = tf.name
            try:
                if "zara.com" in url and "zarahome" not in url: hint = "zara"
                elif "zarahome.com" in url: hint = "zarahome"
                elif "falabella.com" in url: hint = "falabella"
                else: hint = "0"
                result = subprocess.run(
                    [sys.executable, script_path, url, hint],
                    capture_output=True, text=True, timeout=75
                )
                output = result.stdout.strip()
                if not output:
                    logger.error(f"[Playwright] sin output: {result.stderr[:300]}")
                    return ResultadoScrape(url=url, estado="error", empresa=empresa)
                data = _json.loads(output)
            finally:
                os.unlink(script_path)
        precio_num = data.get("precio")
        nombre     = data.get("nombre")
        imagen     = data.get("imagen")
        precio_txt = ("$" + fmt(precio_num)) if precio_num else None
        return ResultadoScrape(url=url, nombre=nombre, precio_txt=precio_txt,
                               precio_num=precio_num, imagen=imagen,
                               estado="ok" if precio_num else "sin_precio", empresa=empresa)
    except subprocess.TimeoutExpired:
        return ResultadoScrape(url=url, estado="error", empresa=empresa)
    except Exception as e:
        logger.error(f"[Playwright] {e} <- {url[:70]}")
        return ResultadoScrape(url=url, estado="error", empresa=empresa)

def _extraer_imagen(soup, url) -> Optional[str]:
    if "paris.cl" in url:
        for img in soup.find_all("img"):
            src = img.get("src", "")
            if "cl-cenco-pim-resizer" in src or "cl-dam-resizer" in src:
                return src
    meta = soup.find("meta", property="og:image")
    if meta and meta.get("content"): return meta["content"]
    for img in soup.find_all("img"):
        c = img.get("data-large_image") or img.get("data-src") or img.get("src")
        if c and not any(x in c.lower() for x in ["logo","icon","banner","sprite"]):
            return c
    return None

# ─────────────────────────────────────────────
# SCRAPER
# ─────────────────────────────────────────────
@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def _scrape_con_requests_cached(url: str) -> ResultadoScrape:
    empresa = get_empresa(url)
    try:
        r = _hacer_request(url)
        if r.status_code in (403, 406, 429, 503):
            return ResultadoScrape(url=url, estado="bloqueado", empresa=empresa)
        soup = BeautifulSoup(r.text, "html.parser")
        og = soup.find("meta", property="og:title")
        nombre = limpiar_texto(og["content"]) if og and og.get("content") else (
            limpiar_texto(soup.title.text) if soup.title else "Sin nombre"
        )
        precio_num = None
        for nombre_est, fn in ESTRATEGIAS:
            try:
                precio_num = fn(soup=soup, html=r.text, url=url)
                if precio_num:
                    logger.info(f"[{nombre_est}] ${precio_num} <- {url[:60]}")
                    break
            except Exception as e:
                logger.warning(f"[{nombre_est}] {e}")
        precio_txt = ("$" + fmt(precio_num)) if precio_num else None
        imagen     = _extraer_imagen(soup, url)
        return ResultadoScrape(url=url, nombre=nombre, precio_txt=precio_txt,
                               precio_num=precio_num, imagen=imagen, estado="ok", empresa=empresa)
    except requests.exceptions.SSLError:
        try:
            import urllib3; urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT, verify=False)
            soup = BeautifulSoup(r.text, "html.parser")
            og = soup.find("meta", property="og:title")
            nombre = limpiar_texto(og["content"]) if og and og.get("content") else (
                limpiar_texto(soup.title.text) if soup.title else "Sin nombre"
            )
            precio_num = None
            for nombre_est, fn in ESTRATEGIAS:
                try:
                    precio_num = fn(soup=soup, html=r.text, url=url)
                    if precio_num: break
                except: pass
            precio_txt = ("$" + fmt(precio_num)) if precio_num else None
            imagen     = _extraer_imagen(soup, url)
            return ResultadoScrape(url=url, nombre=nombre, precio_txt=precio_txt,
                                   precio_num=precio_num, imagen=imagen, estado="ok", empresa=empresa)
        except Exception as e2:
            logger.error(f"[SSL retry fallido] {url}: {e2}")
            return ResultadoScrape(url=url, estado="error", empresa=empresa)
    except Exception as e:
        logger.error(f"Error {url}: {e}")
        return ResultadoScrape(url=url, estado="error", empresa=empresa)

def scrape_producto(url: str) -> ResultadoScrape:
    if not url or (isinstance(url, float) and pd.isna(url)):
        return ResultadoScrape(url="", estado="error")
    if any(d in url for d in PLAYWRIGHT_DOMAINS):
        return _scrape_playwright(url, get_empresa(url))
    return _scrape_con_requests_cached(url)

def scrape_paralelo(urls, progress_cb=None) -> dict:
    res = {}; done = 0; total = len(urls)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        fut = {ex.submit(scrape_producto, u): u for u in urls if u}
        for f in as_completed(fut):
            u = fut[f]
            try: res[u] = f.result()
            except: res[u] = ResultadoScrape(url=u, estado="error")
            done += 1
            if progress_cb: progress_cb(done / total)
    return res

# ─────────────────────────────────────────────
# RENDER HELPERS
# ─────────────────────────────────────────────
def r_img(url_img, url_link=None, size=52):
    if not url_img: return '<span style="color:#1e2d47">—</span>'
    img = f'<img src="{url_img}" width="{size}" height="{size}" class="img-thumb">'
    return f'<a href="{url_link}" target="_blank">{img}</a>' if url_link else img
def r_nombre(nombre, url, maxlen=50):
    if not nombre: return '<span class="status-none">—</span>'
    short = (nombre[:maxlen]+"…") if len(nombre or "")>maxlen else nombre
    if url: return f'<span class="product-name"><a href="{url}" target="_blank" title="{nombre}">{short}</a></span>'
    return f'<span class="product-name">{short}</span>'
def r_precio(txt, nuestro=False):
    if not txt: return '<span class="status-none">sin precio</span>'
    cls = "price-pill price-our" if nuestro else "price-pill"
    return f'<span class="{cls}">{txt}</span>'
def r_dif(dif):
    if dif is None: return '<span class="dif-igual">—</span>'
    if dif > 0:     return f'<span class="dif-mas">▲ +{fmt(dif)}</span>'
    if dif < 0:     return f'<span class="dif-menos">▼ -{fmt(abs(dif))}</span>'
    return '<span class="dif-igual">= 0</span>'
def r_cambio(precio_actual: Optional[int], precio_prev: Optional[int]) -> str:
    if precio_actual is None or precio_prev is None: return ""
    if precio_actual > precio_prev: return f'<span class="price-up">↑ subió ${fmt(precio_actual - precio_prev)}</span>'
    if precio_actual < precio_prev: return f'<span class="price-down">↓ bajó ${fmt(precio_prev - precio_actual)}</span>'
    return ""
def _r_margen(margen_pct, precio_form: Optional[int]) -> str:
    if precio_form is None or margen_pct is None or margen_pct == "":
        return '<span class="status-none">—</span>'
    try:
        m = float(margen_pct)
        if m > 1: m = m / 100
        monto = int(precio_form * m)
        return (f'<span class="price-pill">${fmt(monto)}'
                f'<br><span style="font-size:0.68rem;color:#5a5a5a">{m*100:.1f}%</span></span>')
    except:
        return '<span class="status-none">—</span>'

# ─────────────────────────────────────────────
# REPORTE EXCEL
# ─────────────────────────────────────────────
_Y="F5D000"; _BK="111111"; _W="FFFFFF"; _GL="F2F2F0"; _GN="1A9E6E"; _RD="D93A3A"; _GM="5A5A5A"

def _xfill(c): return PatternFill("solid", start_color=c, end_color=c)
def _xfont(bold=False, color=_BK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
def _xborder(color="E4E4E4"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _xalign(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _xcw(ws, d):
    for col, w in d.items(): ws.column_dimensions[col].width = w
def _xrh(ws, row, h): ws.row_dimensions[row].height = h
def _xstrip(html):
    if not html or not isinstance(html, str): return str(html) if html else "—"
    text = re.sub(r"<[^>]+>", "", str(html))
    return text.replace("&amp;","&").replace("&lt;","<").replace("&gt;",">").strip() or "—"

def _rpt_tabla(wb, df):
    ws = wb.active; ws.title = "Tabla Completa"; ws.sheet_view.showGridLines = False
    hdrs = ["SKU","Rubro","Producto","Precio Nuestro","Tienda","Prod. Rival","Precio Rival","Dif $","Dif %","Margen %","Estado"]
    r = 1
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _xfont(True, _BK, 9); c.fill = _xfill(_Y)
        c.alignment = _xalign("center"); c.border = _xborder(_Y)
    _xrh(ws, r, 22)
    for i, (_, row) in enumerate(df.iterrows()):
        r += 1; bg = _GL if i % 2 == 0 else _W
        pf  = row.get("_precio_form"); pc = row.get("_precio_comp"); dn = row.get("_dif_num")
        dp  = round(dn / pc * 100, 1) if (dn and pc) else None
        margen_pct_val = None
        try:
            raw_m = row.get("_row_excel", {}).get("Margen") if row.get("_row_excel") else None
            if raw_m is not None and raw_m != "":
                m = float(raw_m)
                if m > 1: m = m / 100
                margen_pct_val = m
        except: pass
        if pc is None:          estado = "Sin precio"
        elif dn and dn > 0:     estado = "Sobre precio competencia"
        elif dn and dn < 0:     estado = "Bajo precio competencia"
        else:                   estado = "Igual"
        nombre_prod = _xstrip(str(row.get("_nombre","")))
        nombre_comp = _xstrip(str(row.get("Producto Comp.","")))
        url_form    = str(row.get("_url_form","")) if row.get("_url_form") else ""
        url_comp    = str(row.get("_url_comp","")) if row.get("_url_comp") else ""
        vals = [
            row.get("_sku",""), _xstrip(str(row.get("_rubro",""))), nombre_prod,
            pf or "—", row.get("_empresa",""), nombre_comp, pc or "—",
            dn if dn is not None else "—",
            (dp/100) if dp is not None else "—",
            margen_pct_val if margen_pct_val is not None else "—", estado,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = _xfill(bg); c.border = _xborder()
            c.alignment = _xalign("center" if ci in (4,7,8,9,10) else "left")
            if ci == 8 and isinstance(v,(int,float)):
                c.font = _xfont(bold=True, size=8, color=_GN if v>0 else _RD if v<0 else _GM)
                c.number_format = '$#,##0'
            elif ci == 9 and isinstance(v,(int,float)):
                c.font = _xfont(bold=True, size=8, color=_GN if v>0 else _RD if v<0 else _GM)
                c.number_format = '0.0%'
            elif ci == 10 and isinstance(v, float):
                c.font = _xfont(size=8); c.number_format = '0.0%'
            elif ci in (4,7) and isinstance(v,(int,float)):
                c.font = _xfont(size=8); c.number_format = '$#,##0'
            elif ci == 11:
                c.font = _xfont(size=8, bold=True,
                    color=_GN if v=="Sobre precio competencia" else _RD if v=="Bajo precio competencia" else _GM)
            else:
                c.font = _xfont(size=8)
            if ci == 3 and url_form:
                c.hyperlink = url_form; c.font = _xfont(size=8, color="0563C1"); c.alignment = _xalign("left")
            elif ci == 6 and url_comp:
                c.hyperlink = url_comp; c.font = _xfont(size=8, color="0563C1"); c.alignment = _xalign("left")
        _xrh(ws, r, 15)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:K{r}"
    _xcw(ws, {"A":10,"B":13,"C":28,"D":13,"E":13,"F":26,"G":13,"H":13,"I":8,"J":9,"K":12})

def _rpt_resumen_sku(wb, df, df_excel):
    ws = wb.create_sheet("Resumen por SKU"); ws.sheet_view.showGridLines = False
    hdrs = ["SKU","Nombre","Rubro","Precio Nuestro","Precio Prom. Comp.",
            "vs Promedio $","vs Promedio %","Rival más barato","Precio rival min",
            "N° rivales","Margen %","Cobertura","Posición"]
    r = 1
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = _xfont(True, _BK, 9); c.fill = _xfill(_Y)
        c.alignment = _xalign("center"); c.border = _xborder(_Y)
    _xrh(ws, r, 22)
    for i, (sku, grp) in enumerate(df.groupby("_sku")):
        r += 1; bg = _GL if i % 2 == 0 else _W
        nombre  = _xstrip(str(grp["_nombre"].iloc[0]))
        rubro   = _xstrip(str(grp["_rubro"].iloc[0]))
        pf      = grp["_precio_form"].dropna().iloc[0] if grp["_precio_form"].notna().any() else None
        precios_comp = [float(p) for p in grp["_precio_comp"].dropna().tolist() if p]
        prom_comp  = int(sum(precios_comp)/len(precios_comp)) if precios_comp else None
        min_comp   = min(precios_comp) if precios_comp else None
        rival_min  = "—"
        if min_comp:
            fila_min = grp[grp["_precio_comp"] == min_comp]
            if not fila_min.empty: rival_min = fila_min["_empresa"].iloc[0]
        vs_prom     = (pf - prom_comp) if pf and prom_comp else None
        vs_prom_pct = (vs_prom / prom_comp) if vs_prom and prom_comp else None
        margen_pct_val = None; cobertura_val = None
        rows_ex = df_excel[df_excel["SKU"].astype(str) == str(sku)]
        if not rows_ex.empty:
            try:
                raw_m = rows_ex.iloc[0].get("Margen")
                if raw_m is not None and str(raw_m).strip() != "":
                    m = float(raw_m); m = m/100 if m > 1 else m
                    margen_pct_val = m
            except: pass
            try:
                cv = rows_ex.iloc[0].get("Cobertura")
                if cv is not None and str(cv).strip() != "": cobertura_val = float(cv)
            except: pass
        if not precios_comp:    posicion = "Sin datos"
        elif not pf:            posicion = "Sin precio Form"
        elif pf <= min_comp:    posicion = "Más barato"
        elif pf <= prom_comp:   posicion = "Bajo promedio"
        else:                   posicion = "Sobre promedio"
        vals = [sku, nombre, rubro, pf or "—", prom_comp or "—",
                vs_prom if vs_prom is not None else "—",
                vs_prom_pct if vs_prom_pct is not None else "—",
                rival_min, min_comp or "—", len(precios_comp),
                margen_pct_val if margen_pct_val is not None else "—",
                cobertura_val if cobertura_val is not None else "—", posicion]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = _xfill(bg); c.border = _xborder()
            c.alignment = _xalign("center" if ci not in (2,3,8) else "left")
            if ci in (4,5,9) and isinstance(v,(int,float)):
                c.font = _xfont(size=8); c.number_format = '$#,##0'
            elif ci == 6 and isinstance(v,(int,float)):
                c.font = _xfont(bold=True, size=8, color=_GN if v>0 else _RD if v<0 else _GM)
                c.number_format = '$#,##0'
            elif ci in (7,11) and isinstance(v, float):
                c.font = _xfont(size=8); c.number_format = '0.0%'
            elif ci == 13:
                c.font = _xfont(bold=True, size=8,
                    color=_GN if v=="Más barato" else _RD if v=="Sobre promedio" else _GM)
            else:
                c.font = _xfont(size=8)
        _xrh(ws, r, 15)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:M{r}"
    _xcw(ws, {"A":10,"B":28,"C":14,"D":13,"E":14,"F":13,"G":9,"H":16,"I":13,"J":8,"K":9,"L":10,"M":14})

def generar_reporte(df_final: pd.DataFrame, df_excel: pd.DataFrame) -> bytes:
    wb = Workbook()
    _rpt_tabla(wb, df_final)
    _rpt_resumen_sku(wb, df_final, df_excel)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def _aplicar_precios_editados_a_df(df_base: pd.DataFrame, precios_editados: dict) -> pd.DataFrame:
    df_out = df_base.copy()
    for idx, row in df_out.iterrows():
        key = (str(row.get("_sku", "")).strip(), str(row.get("_empresa", "")).strip())
        if key in precios_editados:
            nuevo_pc = precios_editados[key]
            pf = row.get("_precio_form")
            try:
                pf = int(pf) if pf not in (None, "", "nan") and str(pf) != "nan" else None
            except:
                pf = None
            dif_nuevo = (pf - nuevo_pc) if pf and nuevo_pc else None
            df_out.at[idx, "_precio_comp"] = nuevo_pc
            df_out.at[idx, "_dif_num"] = dif_nuevo
            if "Precio Comp." in df_out.columns:
                df_out.at[idx, "Precio Comp."] = (
                    f'<span class="price-pill price-editado">${fmt(nuevo_pc)}'
                    f'<span class="editado-tag">✎ editado</span></span>'
                )
            if "Diferencia" in df_out.columns:
                df_out.at[idx, "Diferencia"] = r_dif(dif_nuevo)
    return df_out

def _generar_excel_con_ediciones() -> bytes:
    df_base = st.session_state["df_final"].copy()
    precios_editados = st.session_state.get("precios_editados", {})
    df_rep = _aplicar_precios_editados_a_df(df_base, precios_editados)
    return generar_reporte(df_rep, df)

# ─────────────────────────────────────────────
# MODAL DE ANÁLISIS SKU
# ─────────────────────────────────────────────
def render_modal_sku(sku: str, filas_sku: list[dict], df_excel_row):
    nombre     = df_excel_row.get("Nombre", sku)
    rubro      = df_excel_row.get("Rubro", "—")
    margen_pct = df_excel_row.get("Margen", None)
    cobertura  = df_excel_row.get("Cobertura", None)
    precio_form_original = next((f["_precio_form"] for f in filas_sku if f["_precio_form"]), None)

    sim_key = f"sim_precio_{sku}"
    if sim_key not in st.session_state:
        st.session_state[sim_key] = precio_form_original

    def _es_precio_valido(v):
        try: return v is not None and v != "" and float(v) > 0
        except: return False

    precios_comp  = [float(f["_precio_comp"]) for f in filas_sku if _es_precio_valido(f["_precio_comp"])]
    empresas_comp = [(f["_empresa"], float(f["_precio_comp"])) for f in filas_sku if _es_precio_valido(f["_precio_comp"])]
    prom_comp = int(sum(precios_comp)/len(precios_comp)) if precios_comp else None
    min_comp  = min(precios_comp) if precios_comp else None
    max_comp  = max(precios_comp) if precios_comp else None
    rival_min = next((e for e,p in empresas_comp if p == min_comp), "—")

    precio_activo = st.session_state[sim_key] or precio_form_original
    es_simulado   = (precio_activo != precio_form_original) and precio_form_original
    vs_prom = (precio_activo - prom_comp) if precio_activo and prom_comp else None

    margen_txt = "—"; margen_sim_txt = None
    if margen_pct is not None:
        try:
            m = float(margen_pct)
            if m > 1: m = m / 100
            _p_act = int(precio_activo) if precio_activo else None
            _p_ori = int(precio_form_original) if precio_form_original else None
            if _p_ori and 0 < m < 1:
                costo_derivado = _p_ori * (1 - m)
                monto_ori = int(_p_ori - costo_derivado)
                margen_txt = f"${fmt(monto_ori)} ({m*100:.1f}%)"
                if _p_act and _p_act != _p_ori:
                    m_nuevo = (_p_act - costo_derivado) / _p_act
                    monto_sim = int(_p_act - costo_derivado)
                    margen_sim_txt = f"${fmt(monto_sim)} ({m_nuevo*100:.1f}%)"
        except Exception as e: logger.debug(f"[margen] {e}")

    precio_badge = ""
    if es_simulado:
        precio_badge = ' <span style="font-size:0.6rem;background:#4b7cff;color:#fff;padding:2px 7px;border-radius:20px;vertical-align:middle;margin-left:6px">✎ simulado</span>'

    st.markdown(f"""
    <div class="modal-title">{nombre}</div>
    <div class="modal-sub">SKU <span style="font-family:'DM Mono',monospace;color:#4b7cff">{sku}</span>
     · <span class="rubro-tag">{rubro}</span></div>
    <div class="kpi-grid">
        <div class="kpi-card" style="{'border:2px solid #4b7cff;' if es_simulado else ''}">
            <div class="kpi-label">Precio Nuestro{precio_badge}</div>
            <div class="kpi-val" style="color:#111;background:{'#dce8ff' if es_simulado else '#F5D000'};display:inline-block;padding:2px 10px;border-radius:5px">${fmt(precio_activo) if precio_activo else "—"}</div>
            <div class="kpi-sub">{'precio original: $' + fmt(precio_form_original) if es_simulado and precio_form_original else 'precio web actual'}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Precio Promedio Comp.</div>
            <div class="kpi-val">${fmt(prom_comp) if prom_comp else "—"}</div>
            <div class="kpi-sub">{len(precios_comp)} rivales con precio</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">vs. Promedio Comp.</div>
            <div class="kpi-val" style="color:{'#ff4d6a' if vs_prom and vs_prom>0 else '#00d68f' if vs_prom and vs_prom<0 else '#f0f4ff'}">
                {'▲' if vs_prom and vs_prom>0 else '▼' if vs_prom and vs_prom<0 else '='} ${fmt(abs(vs_prom)) if vs_prom else "—"}
            </div>
            <div class="kpi-sub">{'más caros' if vs_prom and vs_prom>0 else 'más baratos' if vs_prom and vs_prom<0 else 'igual al promedio'}</div>
        </div>
        <div class="kpi-card" style="{'border:2px solid #4b7cff;' if margen_sim_txt else ''}">
            <div class="kpi-label">Margen{'<span style="font-size:0.6rem;background:#4b7cff;color:#fff;padding:2px 6px;border-radius:20px;margin-left:5px">simulado</span>' if margen_sim_txt else ''}</div>
            <div class="kpi-val" style="font-size:1.1rem">{margen_sim_txt if margen_sim_txt else margen_txt}</div>
            <div class="kpi-sub">{'original: ' + margen_txt if margen_sim_txt else 'sobre precio web'}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Cobertura</div>
            <div class="kpi-val" style="font-size:1.1rem">{f"{float(cobertura):.2f}" if cobertura is not None and cobertura != "" else "—"}</div>
            <div class="kpi-sub">unidades disponibles</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-label">Rival más barato</div>
            <div class="kpi-val" style="font-size:1.05rem;color:#1a9e6e">{rival_min}</div>
            <div class="kpi-sub">${fmt(min_comp) if min_comp else "—"}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _p_sim_actual = int(st.session_state[sim_key]) if st.session_state[sim_key] else (int(precio_form_original) if precio_form_original else 10000)
    _p_ori_int    = int(precio_form_original) if precio_form_original else None
    _es_original  = (_p_sim_actual == _p_ori_int) if _p_ori_int else True

    st.markdown('<div class="section-title">💡 Simular Precio</div>', unsafe_allow_html=True)

    def _on_sim_change():
        st.session_state[sim_key] = st.session_state[f"sim_input_{sku}"]

    st.number_input(
        "Nuevo precio a simular",
        min_value=1000, max_value=10_000_000,
        value=_p_sim_actual, step=1000,
        key=f"sim_input_{sku}", format="%d",
        on_change=_on_sim_change,
    )
    if not _es_original:
        if st.button(f"↩ Volver al precio original  ${fmt(_p_ori_int)}", key=f"sim_reset_{sku}", type="secondary"):
            st.session_state[sim_key] = precio_form_original
            if f"sim_input_{sku}" in st.session_state:
                del st.session_state[f"sim_input_{sku}"]
            st.rerun()

    st.markdown('<div class="section-title">Detalle por Rival</div>', unsafe_allow_html=True)
    for e, p in sorted(empresas_comp, key=lambda x: x[1]):
        dif = (precio_activo - p) if precio_activo and p else None
        bar_pct = int(p / max_comp * 100) if max_comp else 50
        bar_col = "#ff4d6a" if dif and dif > 0 else "#00d68f" if dif and dif < 0 else "#3b6bff"
        st.markdown(f"""
        <div class="rival-row">
            <span class="rival-name">{e}</span>
            <div style="flex:1;margin:0 16px">
                <div style="background:#131927;border-radius:4px;height:4px;overflow:hidden">
                    <div style="width:{bar_pct}%;background:{bar_col};height:100%;border-radius:4px"></div>
                </div>
            </div>
            <span class="rival-price">${fmt(p)}</span>
            <span style="margin-left:14px;min-width:80px;text-align:right">{r_dif(dif)}</span>
        </div>
        """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CARGA EXCEL BASE
# ─────────────────────────────────────────────
try:
    df = pd.read_excel(RUTA_EXCEL)
except FileNotFoundError:
    st.error("❌ No se encontró `base_precios.xlsx`."); st.stop()
columnas_comp = [c for c in df.columns if str(c).startswith("Link_Comp")]
rubros_disponibles = sorted(df["Rubro"].dropna().unique().tolist()) if "Rubro" in df.columns else []

# ─────────────────────────────────────────────
# PRECIOS EDITADOS — cargar desde Supabase al inicio
# ─────────────────────────────────────────────
if "precios_editados" not in st.session_state:
    st.session_state["precios_editados"] = _cargar_precios_editados_sb()

# ─────────────────────────────────────────────
# RECEPTOR DE EDICIONES DE PRECIO (query_params)
# ─────────────────────────────────────────────
_qp = st.query_params
if "reset_sku" in _qp and "reset_emp" in _qp:
    try:
        _sku_r = str(_qp["reset_sku"]).strip()
        _emp_r = str(_qp["reset_emp"]).strip()
        st.session_state["precios_editados"].pop((_sku_r, _emp_r), None)
        _eliminar_precio_editado_sb(_sku_r, _emp_r)
        st.query_params.clear()
        st.rerun()
    except: pass

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
ts_str = st.session_state.get("timestamp", "")
badge  = f'<span class="header-badge">↻ {ts_str}</span>' if ts_str else ""
st.markdown(f"""
<div class="header-wrap">
    <div class="header-left">
        <div class="header-accent"></div>
        <div>
            <div class="header-title">Monitor de Precios</div>
            <div class="header-sub">Sistema Interno de Monitoreo Competitivo · Form Design</div>
        </div>
    </div>
    {badge}
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONTROLES
# ─────────────────────────────────────────────
if MODO_VISTA:
    if "df_final" not in st.session_state:
        _df_snap, _ts_snap = cargar_snapshot()
        if _df_snap is not None:
            st.session_state["df_final"]  = _df_snap
            st.session_state["timestamp"] = _ts_snap
            st.session_state["sku_modal"] = None
    ts_snap = st.session_state.get("timestamp", "—")
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:12px;background:#fdf9e3;border:1.5px solid #F5D000;
                border-radius:10px;padding:10px 18px;margin-bottom:16px">
        <span style="font-size:1.3rem">👁️</span>
        <div>
            <span style="font-family:'Syne',sans-serif;font-weight:800;font-size:0.88rem;color:#111">
                Modo Vista — datos del último scrape
            </span>
            <span style="font-size:0.75rem;color:#5a5a5a;margin-left:10px">
                Actualizado: <b>{ts_snap}</b>
            </span>
        </div>
    </div>
    """, unsafe_allow_html=True)
    col_rubro, col_emp, col_dif = st.columns([2, 2, 2])
    actualizar = False
else:
    col_btn, col_rubro, col_emp, col_dif = st.columns([1.2, 2, 2, 2])
    with col_btn:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        actualizar = st.button("⟳  Actualizar Precios", use_container_width=True)

with col_rubro:
    filtro_rubro = st.multiselect("Rubro", rubros_disponibles, placeholder="Todos los rubros")
with col_emp:
    todas_empresas = sorted({
        get_empresa(str(row[col]))
        for _, row in df.iterrows()
        for col in columnas_comp
        if pd.notna(row.get(col))
    })
    filtro_empresa = st.multiselect("Tienda", todas_empresas, placeholder="Todas las tiendas")
with col_dif:
    filtro_dif = st.selectbox("Diferencia", ["Todos","Nosotros más caros que la competencia","Nosotros más baratos que la competencia","Igual"])
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# PROCESO DE SCRAPE
# ─────────────────────────────────────────────
if actualizar:
    _scrape_con_requests_cached.clear()
    todas_urls = set()
    for _, row in df.iterrows():
        if pd.notna(row.get("Link_Nuestro")): todas_urls.add(str(row["Link_Nuestro"]))
        for col in columnas_comp:
            if pd.notna(row.get(col)): todas_urls.add(str(row[col]))
    barra = st.progress(0, text="Scrapeando en paralelo…")
    cache_res = scrape_paralelo(list(todas_urls), lambda p: barra.progress(p, text=f"Scrapeando… {int(p*100)}%"))
    barra.progress(1.0, text="¡Listo!")
    filas = []; filas_db = []
    for _, row in df.iterrows():
        sku       = limpiar_texto(row.get("SKU",""))
        nom_excel = limpiar_texto(row.get("Nombre",""))
        rubro     = limpiar_texto(row.get("Rubro",""))
        url_form  = str(row.get("Link_Nuestro","")) if pd.notna(row.get("Link_Nuestro")) else ""
        res_form  = cache_res.get(url_form, ResultadoScrape(url=url_form))
        filas_db.append({"sku":sku,"url":url_form,"empresa":"Nosotros","precio":res_form.precio_num,"estado":res_form.estado})
        for col in columnas_comp:
            url_comp = str(row.get(col,"")) if pd.notna(row.get(col)) else ""
            if not url_comp: continue
            res_comp    = cache_res.get(url_comp, ResultadoScrape(url=url_comp))
            prev        = precio_anterior(sku, url_comp)
            dif         = ((res_form.precio_num - res_comp.precio_num) if res_form.precio_num and res_comp.precio_num else None)
            cambio_html = r_cambio(res_comp.precio_num, prev)
            precio_comp_html = r_precio(res_comp.precio_txt) + (f"<br>{cambio_html}" if cambio_html else "")
            filas_db.append({"sku":sku,"url":url_comp,"empresa":res_comp.empresa,"precio":res_comp.precio_num,"estado":res_comp.estado})
            filas.append({
                "_sku": sku, "_nombre": nom_excel or res_form.nombre or sku,
                "_empresa": res_comp.empresa, "_rubro": rubro, "_dif_num": dif,
                "_precio_form": res_form.precio_num, "_precio_comp": res_comp.precio_num,
                "_url_form": url_form, "_url_comp": url_comp, "_row_excel": row.to_dict(),
                "_busqueda": f"{sku} {nom_excel or ''} {res_form.nombre or ''} {res_comp.nombre or ''} {rubro or ''}".lower(),
                "SKU":            f'<span class="sku-btn" data-sku="{sku}">{sku}</span>' if sku else "",
                "Rubro":          f'<span class="rubro-tag">{rubro}</span>' if rubro else "",
                "Foto":           r_img(res_form.imagen, url_form, 48),
                "Producto":       r_nombre(nom_excel or res_form.nombre, url_form),
                "Precio Nuestro": r_precio(res_form.precio_txt, nuestro=True),
                "Tienda":         f'<span class="empresa-tag">{res_comp.empresa}</span>',
                "Foto Comp.":     r_img(res_comp.imagen, url_comp, 48),
                "Producto Comp.": r_nombre(res_comp.nombre, url_comp),
                "Precio Comp.":   precio_comp_html,
                "Diferencia":     r_dif(dif),
                "Margen":         _r_margen(row.get("Margen"), res_form.precio_num),
            })
    guardar_precios(filas_db)
    _df_nuevo = pd.DataFrame(filas)
    _ts_nuevo = time.strftime("%d/%m/%Y %H:%M:%S")

    _pe = st.session_state.get("precios_editados", {})
    if _pe:
        for idx_r, row in _df_nuevo.iterrows():
            k = (str(row.get("_sku","")), str(row.get("_empresa","")))
            if k in _pe:
                nuevo_pc  = _pe[k]
                pf        = row.get("_precio_form")
                dif_nueva = (pf - nuevo_pc) if pf and nuevo_pc else None
                _df_nuevo.at[idx_r, "_precio_comp"] = nuevo_pc
                _df_nuevo.at[idx_r, "_dif_num"]     = dif_nueva
                _df_nuevo.at[idx_r, "Precio Comp."] = (
                    f'<span class="price-pill price-editado">${fmt(nuevo_pc)}'
                    f'<span class="editado-tag">✎ editado</span></span>'
                )
                _df_nuevo.at[idx_r, "Diferencia"] = r_dif(dif_nueva)

    st.session_state["df_final"]  = _df_nuevo
    st.session_state["timestamp"] = _ts_nuevo
    st.session_state["sku_modal"] = None
    guardar_snapshot(_df_nuevo, _ts_nuevo)
    st.rerun()

# ─────────────────────────────────────────────
# MOSTRAR TABLA
# ─────────────────────────────────────────────
if "df_final" in st.session_state:
    df_vis = st.session_state["df_final"].copy()
    if filtro_rubro:    df_vis = df_vis[df_vis["_rubro"].isin(filtro_rubro)]
    if filtro_empresa:  df_vis = df_vis[df_vis["_empresa"].isin(filtro_empresa)]
    if filtro_dif == "Nosotros más caros que la competencia":     df_vis = df_vis[df_vis["_dif_num"].notna() & (df_vis["_dif_num"] > 0)]
    elif filtro_dif == "Nosotros más baratos que la competencia": df_vis = df_vis[df_vis["_dif_num"].notna() & (df_vis["_dif_num"] < 0)]
    elif filtro_dif == "Igual":            df_vis = df_vis[df_vis["_dif_num"].notna() & (df_vis["_dif_num"] == 0)]

    total     = len(df_vis)
    mas_caros = int((df_vis["_dif_num"] > 0).sum())
    mas_bar   = int((df_vis["_dif_num"] < 0).sum())
    sin_p     = int(df_vis["_precio_comp"].isna().sum())
    st.markdown(f"""
    <div class="metrics-row">
        <div class="metric-card yellow"><div class="metric-label">Comparaciones</div><div class="metric-value">{total}</div><div class="metric-sub">filas visibles</div></div>
        <div class="metric-card green"><div class="metric-label">Más caros que la competencia</div><div class="metric-value" style="color:#1a9e6e">{mas_caros}</div><div class="metric-sub">donde ganamos en precio</div></div>
        <div class="metric-card red"><div class="metric-label">Más baratos que la competencia</div><div class="metric-value" style="color:#d93a3a">{mas_bar}</div><div class="metric-sub">donde nos superan</div></div>
        <div class="metric-card gray"><div class="metric-label">Sin precio</div><div class="metric-value" style="color:#aaa">{sin_p}</div><div class="metric-sub">no scrapeados</div></div>
    </div>
    """, unsafe_allow_html=True)

    import streamlit.components.v1 as components

    _precios_ed = st.session_state.get("precios_editados", {})
    for _idx_v, _row_v in df_vis.iterrows():
        _key_v = (str(_row_v.get("_sku","")), str(_row_v.get("_empresa","")))
        if _key_v in _precios_ed:
            _pc_v = _precios_ed[_key_v]
            _pf_v = _row_v.get("_precio_form")
            _dif_v = (_pf_v - _pc_v) if _pf_v and _pc_v else None
            df_vis.at[_idx_v, "_precio_comp"] = _pc_v
            df_vis.at[_idx_v, "_dif_num"]     = _dif_v
            df_vis.at[_idx_v, "Precio Comp."] = (
                f'<span class="price-pill price-editado">${fmt(_pc_v)}'
                f'<span class="editado-tag">✎ editado</span></span>'
            )
            df_vis.at[_idx_v, "Diferencia"] = r_dif(_dif_v)

    sku_nombre_map   = df_vis[["_sku","_nombre"]].drop_duplicates("_sku").set_index("_sku")["_nombre"].to_dict()
    opciones_display = {f"{v}  ({k})": k for k, v in sku_nombre_map.items()}
    opciones_lista   = ["— ninguno —"] + sorted(opciones_display.keys())
    sel_display = st.selectbox("📊 Seleccionar producto para análisis (escribe para buscar)", opciones_lista, index=0, key="sku_selectbox")
    sku_activo  = None if sel_display == "— ninguno —" else opciones_display.get(sel_display)
    if sku_activo and sku_activo in df_vis["_sku"].values:
        filas_sku  = df_vis[df_vis["_sku"] == sku_activo].to_dict("records")
        rows_match = df[df["SKU"].astype(str) == str(sku_activo)]
        row_excel  = rows_match.iloc[0].to_dict() if len(rows_match) > 0 else {}
        with st.expander(f"📊 {sku_nombre_map.get(sku_activo, sku_activo)}", expanded=True):
            render_modal_sku(sku_activo, filas_sku, row_excel)

    _df_excel = st.session_state["df_final"].copy()
    for _idx_e, _row_e in _df_excel.iterrows():
        _key_e = (str(_row_e.get("_sku","")), str(_row_e.get("_empresa","")))
        if _key_e in _precios_ed:
            _pc_e = _precios_ed[_key_e]
            _pf_e = _row_e.get("_precio_form")
            try:
                _pf_e = int(_pf_e) if _pf_e not in (None,"","nan") and str(_pf_e)!="nan" else None
            except:
                _pf_e = None
            _df_excel.at[_idx_e, "_precio_comp"] = _pc_e
            _df_excel.at[_idx_e, "_dif_num"]     = (_pf_e - _pc_e) if _pf_e else None

    _excel_bytes = generar_reporte(_df_excel, df)
    st.download_button(
        label="⬇  Descargar Reporte",
        data=_excel_bytes,
        file_name=f"reporte_precios_{time.strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{len(_precios_ed)}",
    )

    import os as _os
    _comp_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "table_component")
    _tabla_comp = components.declare_component("price_table", path=_comp_path)

    import json as _json
    def _si(v):
        try: f = float(v); return int(f) if f == f else 0
        except: return 0

    _filas_comp = []
    for _fc_fila in df_vis.to_dict("records"):
        _fc_sku = _fc_fila.get("_sku","")
        _fc_emp = _fc_fila.get("_empresa","")
        _fc_key = (_fc_sku, _fc_emp)
        _fc_pc  = _si(_fc_fila.get("_precio_comp"))
        _filas_comp.append({
            "_sku":         _fc_sku,
            "_empresa":     _fc_emp,
            "_nombre":      str(_fc_fila.get("_nombre","")),
            "_rubro":       str(_fc_fila.get("_rubro","")),
            "_precio_form": _si(_fc_fila.get("_precio_form")),
            "_precio_comp": _fc_pc,
            "_editado":     _fc_key in _precios_ed,
            "Rubro":        _fc_fila.get("Rubro",""),
            "Foto":         _fc_fila.get("Foto",""),
            "Producto":     _fc_fila.get("Producto",""),
            "Precio Nuestro": _fc_fila.get("Precio Nuestro",""),
            "Tienda":       _fc_fila.get("Tienda",""),
            "Foto Comp.":   _fc_fila.get("Foto Comp.",""),
            "Producto Comp.": _fc_fila.get("Producto Comp.",""),
            "Precio Comp.": _fc_fila.get("Precio Comp.",""),
            "Diferencia":   _fc_fila.get("Diferencia",""),
            "Margen":       _fc_fila.get("Margen",""),
        })

    n_filas = len(_filas_comp); ROW_H = 48; HEADER_H = 40; SEARCH_H = 48
    altura_scroll = min(520, max(180, HEADER_H + n_filas * ROW_H))
    altura_total  = SEARCH_H + altura_scroll + 4

    _comp_result = _tabla_comp(
        filas=_filas_comp,
        sku_activo=sku_activo or "",
        key=f"tabla_{len(_precios_ed)}",
        default=None,
        height=altura_total,
    )

    if _comp_result and isinstance(_comp_result, dict):
        _action = _comp_result.get("action")
        _r_sku  = str(_comp_result.get("sku","")).strip()
        _r_emp  = str(_comp_result.get("empresa","")).strip()
        if _action == "edit":
            _r_pc = int(_comp_result.get("precio", 0))
            if _r_sku and _r_emp and _r_pc > 0:
                st.session_state["precios_editados"][(_r_sku, _r_emp)] = _r_pc
                _guardar_precio_editado_sb(_r_sku, _r_emp, _r_pc)
                st.rerun()
        elif _action == "reset":
            if _r_sku and _r_emp:
                st.session_state["precios_editados"].pop((_r_sku, _r_emp), None)
                _eliminar_precio_editado_sb(_r_sku, _r_emp)
                st.rerun()

    st.markdown("---")
    st.markdown("#### ✎ Editar precio de competidor")
    _skus_disp  = sorted(df_vis["_sku"].dropna().unique().tolist())
    _col1, _col2, _col3, _col4 = st.columns([2, 2, 2, 1])
    with _col1:
        _e_sku = st.selectbox("SKU", ["— seleccionar —"] + _skus_disp, key="panel_sku")
    with _col2:
        if _e_sku and _e_sku != "— seleccionar —":
            _emps = sorted(df_vis[df_vis["_sku"] == _e_sku]["_empresa"].dropna().unique().tolist())
        else:
            _emps = []
        _e_emp = st.selectbox("Tienda", ["— seleccionar —"] + _emps, key="panel_emp")
    with _col3:
        _e_pc = st.number_input("Precio manual ($)", min_value=0, step=1000, value=0, key="panel_pc")
    with _col4:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("💾 Guardar", key="btn_guardar_precio", use_container_width=True):
            if _e_sku != "— seleccionar —" and _e_emp != "— seleccionar —" and _e_pc > 0:
                _kk = (_e_sku, _e_emp)
                st.session_state["precios_editados"][_kk] = _e_pc
                _guardar_precio_editado_sb(_e_sku, _e_emp, _e_pc)
                st.success(f"✓ Guardado: {_e_sku} / {_e_emp} = ${fmt(_e_pc)}")
                st.rerun()
            else:
                st.warning("Selecciona SKU, tienda e ingresa un precio válido.")

    _precios_ed_panel = st.session_state.get("precios_editados", {})
    if _precios_ed_panel:
        with st.expander(f"✎ Precios editados manualmente ({len(_precios_ed_panel)})", expanded=False):
            _cols_h = st.columns([2,2,2,1])
            _cols_h[0].markdown("**SKU**"); _cols_h[1].markdown("**Tienda**")
            _cols_h[2].markdown("**Precio manual**"); _cols_h[3].markdown("**Acción**")
            for (_sk, _em), _pv in list(_precios_ed_panel.items()):
                _c0,_c1,_c2,_c3 = st.columns([2,2,2,1])
                _c0.write(_sk); _c1.write(_em); _c2.write(f"${fmt(_pv)}")
                if _c3.button("✕", key=f"del_{_sk}_{_em}"):
                    del st.session_state["precios_editados"][(_sk, _em)]
                    _eliminar_precio_editado_sb(_sk, _em)
                    st.rerun()
            if st.button("🗑 Limpiar todos"):
                st.session_state["precios_editados"] = {}
                _limpiar_todos_precios_editados_sb()
                st.rerun()
else:
    if MODO_VISTA:
        st.markdown("""
        <div class="empty-state">
            <div class="empty-icon">⏳</div>
            <div class="empty-text">Aún no hay datos disponibles.<br>
            El responsable del sistema debe ejecutar el scrape primero.</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="empty-state">
            <div class="empty-icon">📊</div>
            <div class="empty-text">Presiona <strong>Actualizar Precios</strong> para comenzar el monitoreo</div>
        </div>
        """, unsafe_allow_html=True)
